import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import RadarChart, Reference, BarChart
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
from datetime import date

# ── Data ──────────────────────────────────────────────────────────────────────

COMPANY = "Janatics India Private Limited"
CIN     = "U31103TZ1991PTC003409"
DATE    = date.today().strftime("%d %B %Y")

dimensions = [
    {
        "code": "SL", "name": "Strategy & Leadership", "weight": 0.25, "score": 81,
        "sub_metrics": [
            ("SL01", "Founder/Company Tenure (est. 1977, 49 yrs)", 90, "Founded 1977; incorporated 1991 — one of India's oldest domestic pneumatics brands"),
            ("SL02", "Director Profile & DIN Compliance", 75, "3 active directors: G Jaganathan, GC Nageswaran, R Ramesh — all DINs active"),
            ("SL03", "Governance & Filing Compliance", 80, "AGM held Sept 30 2024; filings current with ROC Coimbatore"),
            ("SL04", "Debt & Charge Structure", 78, "ICRA-rated; stable outlook; healthy debt protection metrics per ICRA rationale"),
            ("SL05", "Strategic Direction & Diversification", 82, "Expanded beyond pneumatics → electric actuators, robotics, Industry 4.0 alignment"),
            ("SL06", "Global Presence & Export Strategy", 88, "Operations in 42+ countries; subsidiaries in USA and Germany; global distribution network"),
        ]
    },
    {
        "code": "SM", "name": "Sales & Marketing", "weight": 0.20, "score": 78,
        "sub_metrics": [
            ("SM01", "Revenue Growth vs Industry CAGR", 55, "Revenue grew ~4-5% YoY (FY24→FY25) vs industry CAGR of 6.9% — slightly below market"),
            ("SM02", "Revenue Scale", 80, "FY2025 revenue: ₹531 Cr (~$63M USD) — significant scale for domestic manufacturer"),
            ("SM03", "Customer Diversity", 90, "17,000+ customers across pharma, auto, packaging, printing, food, medical, textile"),
            ("SM04", "Distribution Network", 85, "200 distribution partners globally; 3 plants for regional coverage"),
            ("SM05", "Brand Strength", 80, "49-year brand; own-label (Janatics); member of AIA India; DSIR-recognized R&D"),
        ]
    },
    {
        "code": "OSC", "name": "Operations & Supply Chain", "weight": 0.20, "score": 78,
        "sub_metrics": [
            ("OSC01", "Manufacturing Scale & Capacity", 85, "70,000 sqm facility in Coimbatore; additional plants in Ahmedabad and Noida"),
            ("OSC02", "Product Portfolio Breadth", 90, "3,500+ distinct products across pneumatics, vacuum, hydraulics, didactics, robotics"),
            ("OSC03", "R&D & Technology Capability", 82, "DSIR-recognized R&D division; CNC machining; casting operations; expanding to robotics"),
            ("OSC04", "GST & Regulatory Compliance", 72, "Active GST entity; ICRA notes consistent compliance; e-Invoice enablement assumed"),
            ("OSC05", "Supply Chain Digitization", 65, "Industry 4.0 product range signals awareness; internal digitization not publicly quantified"),
        ]
    },
    {
        "code": "FIN", "name": "Finance", "weight": 0.25, "score": 67,
        "sub_metrics": [
            ("FIN01", "Revenue Scale (₹531 Cr FY25)", 80, "₹531 Cr revenue — strong for unlisted domestic manufacturer in pneumatics segment"),
            ("FIN02", "Revenue Growth Rate (4% YoY)", 55, "FY24: ₹507.9 Cr → FY25: ₹527.8 Cr — 4% growth; steady but below industry pace"),
            ("FIN03", "EBITDA Trajectory", 42, "EBITDA CAGR -22% (1-yr) — margin compression flagged; needs monitoring"),
            ("FIN04", "Credit Profile (ICRA Rated)", 82, "ICRA rated with stable outlook; multiple reaffirmations; healthy debt protection metrics"),
            ("FIN05", "Capital Structure", 75, "Auth. capital ₹30 Cr; paid-up ₹26.5 Cr; nearly fully subscribed — conservative but sound"),
            ("FIN06", "Employee Productivity", 70, "₹531 Cr / 662 employees = ₹80L per employee — decent for manufacturing sector"),
        ]
    },
    {
        "code": "IC", "name": "Industry Context", "weight": 0.10, "score": 75,
        "sub_metrics": [
            ("IC01", "India Pneumatics Market Size", 70, "India pneumatics market ~$1.06B (2024); growing to $1.94B by 2033 — mid-size TAM"),
            ("IC02", "Industry Growth Rate (6.9% CAGR)", 78, "India pneumatic equipment CAGR 6.9% (2025–2033); aligned with industrial capex cycle"),
            ("IC03", "Competitive Position", 70, "Domestic brand competing with global leaders (SMC, Festo, Parker); strong in mid-market"),
            ("IC04", "Macro Tailwinds", 82, "Make in India, PLI schemes, manufacturing automation, China+1 strategy all favourable"),
            ("IC05", "End-Market Diversification", 76, "6+ industries: pharma, auto, packaging, food, medical, textile — low single-sector risk"),
        ]
    },
]

# Composite
composite = sum(d["score"] * d["weight"] for d in dimensions)

# ── Style helpers ──────────────────────────────────────────────────────────────

def score_color(score):
    if score >= 75: return "00B050"   # green
    if score >= 50: return "FFC000"   # amber
    return "FF0000"                   # red

def thin_border():
    s = Side(style='thin', color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
AMBER      = "FFC000"
GREEN      = "00B050"
RED_       = "FF0000"

# ── Workbook ───────────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()

# ── Sheet 1: Summary ───────────────────────────────────────────────────────────

ws = wb.active
ws.title = "Scorecard Summary"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 32
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 14
ws.row_dimensions[1].height = 8

# Title block
for r in range(2, 6):
    ws.row_dimensions[r].height = 20
ws.merge_cells("B2:F2")
ws["B2"] = "DDMR — DUE DILIGENCE & MARKET REPORT"
ws["B2"].font = Font(name="Calibri", bold=True, size=16, color=WHITE)
ws["B2"].fill = header_fill(DARK_BLUE)
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("B3:F3")
ws["B3"] = COMPANY
ws["B3"].font = Font(name="Calibri", bold=True, size=13, color=WHITE)
ws["B3"].fill = header_fill(MID_BLUE)
ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("B4:F4")
ws["B4"] = f"CIN: {CIN}   |   Report Date: {DATE}   |   Composite Score: {composite:.1f}/100"
ws["B4"].font = Font(name="Calibri", size=10, color=DARK_BLUE)
ws["B4"].fill = header_fill(LIGHT_BLUE)
ws["B4"].alignment = Alignment(horizontal="center", vertical="center")

# Composite score cell
ws.row_dimensions[5].height = 8
ws.row_dimensions[6].height = 60
ws.merge_cells("B6:C6")
ws["B6"] = f"{composite:.1f}"
ws["B6"].font = Font(name="Calibri", bold=True, size=36, color=WHITE)
ws["B6"].fill = header_fill(score_color(composite))
ws["B6"].alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells("D6:F6")
ws["D6"] = "COMPOSITE DDMR SCORE\n(Weighted across 5 Dimensions)"
ws["D6"].font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
ws["D6"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Column headers
ws.row_dimensions[8].height = 22
for col, val, w in [
    ("B", "Dimension", 32), ("C", "Score /100", 14),
    ("D", "Weight", 14), ("E", "Weighted", 14), ("F", "Rating", 14)
]:
    ws[f"{col}8"] = val
    ws[f"{col}8"].font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    ws[f"{col}8"].fill = header_fill(DARK_BLUE)
    ws[f"{col}8"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"{col}8"].border = thin_border()

# Dimension rows
for i, d in enumerate(dimensions, start=9):
    ws.row_dimensions[i].height = 20
    rating = "Strong" if d["score"] >= 75 else ("Moderate" if d["score"] >= 50 else "Weak")
    vals = [d["name"], d["score"], f"{int(d['weight']*100)}%",
            round(d["score"] * d["weight"], 1), rating]
    for col, val in zip(["B","C","D","E","F"], vals):
        cell = ws[f"{col}{i}"]
        cell.value = val
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        if col == "B":
            cell.alignment = Alignment(horizontal="left", vertical="center")
        if col in ("C", "E"):
            cell.fill = PatternFill("solid", fgColor=score_color(d["score"]) + "40")
        if col == "F":
            color = score_color(d["score"])
            cell.font = Font(name="Calibri", bold=True, size=10, color=color)

# Totals row
tr = 9 + len(dimensions)
ws.row_dimensions[tr].height = 22
ws[f"B{tr}"] = "COMPOSITE SCORE"
ws[f"C{tr}"] = composite
ws[f"D{tr}"] = "100%"
ws[f"E{tr}"] = round(composite, 1)
ws[f"F{tr}"] = "Strong" if composite >= 75 else "Moderate"
for col in ["B","C","D","E","F"]:
    cell = ws[f"{col}{tr}"]
    cell.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    cell.fill = header_fill(MID_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()
ws[f"B{tr}"].alignment = Alignment(horizontal="left", vertical="center")

# Conditional formatting on score column
ws.conditional_formatting.add(
    f"C9:C{tr-1}",
    ColorScaleRule(
        start_type="num", start_value=0,   start_color="FF0000",
        mid_type="num",   mid_value=50,    mid_color="FFFF00",
        end_type="num",   end_value=100,   end_color="00B050"
    )
)

# Radar chart
chart = RadarChart()
chart.type = "filled"
chart.title = "DDMR Dimension Scores"
chart.style = 10
chart.width = 14
chart.height = 12

labels = Reference(ws, min_col=2, min_row=9, max_row=9+len(dimensions)-1)
data   = Reference(ws, min_col=3, min_row=8, max_row=8+len(dimensions))
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)
ws.add_chart(chart, "B15")

# ── Sheets 2–6: Per-dimension ──────────────────────────────────────────────────

for d in dimensions:
    ws2 = wb.create_sheet(title=f"{d['code']} Detail")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 4
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 40
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 50

    # Header
    ws2.row_dimensions[2].height = 24
    ws2.merge_cells("B2:E2")
    ws2["B2"] = f"{d['code']} — {d['name']}   |   Score: {d['score']}/100   |   Weight: {int(d['weight']*100)}%"
    ws2["B2"].font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    ws2["B2"].fill = header_fill(MID_BLUE)
    ws2["B2"].alignment = Alignment(horizontal="center", vertical="center")

    # Col headers
    ws2.row_dimensions[4].height = 20
    for col, val in [("B","Code"),("C","Sub-Metric"),("D","Score /100"),("E","Evidence / Rationale")]:
        ws2[f"{col}4"] = val
        ws2[f"{col}4"].font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        ws2[f"{col}4"].fill = header_fill(DARK_BLUE)
        ws2[f"{col}4"].alignment = Alignment(horizontal="center", vertical="center")
        ws2[f"{col}4"].border = thin_border()

    # Sub-metric rows
    for j, (code, name, score, evidence) in enumerate(d["sub_metrics"], start=5):
        ws2.row_dimensions[j].height = 32
        for col, val in [("B",code),("C",name),("D",score),("E",evidence)]:
            cell = ws2[f"{col}{j}"]
            cell.value = val
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_border()
            if col == "D":
                cell.font = Font(name="Calibri", bold=True, size=11, color=score_color(score))
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == "B":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Calibri", bold=True, size=9, color=MID_BLUE)

    # Avg row
    last = 5 + len(d["sub_metrics"])
    ws2.row_dimensions[last].height = 22
    ws2[f"B{last}"] = "AVG"
    ws2[f"C{last}"] = f"Dimension Average Score"
    ws2[f"D{last}"] = d["score"]
    ws2[f"E{last}"] = f"Weighted contribution to composite: {d['score'] * d['weight']:.1f} pts"
    for col in ["B","C","D","E"]:
        cell = ws2[f"{col}{last}"]
        cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cell.fill = header_fill(MID_BLUE)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2[f"C{last}"].alignment = Alignment(horizontal="left", vertical="center")
    ws2[f"E{last}"].alignment = Alignment(horizontal="left", vertical="center")

    # Bar chart for sub-metrics
    bar = BarChart()
    bar.type = "bar"
    bar.title = f"{d['code']} Sub-Metric Scores"
    bar.style = 10
    bar.width = 18
    bar.height = 10
    sm_labels = Reference(ws2, min_col=2, min_row=5, max_row=5+len(d["sub_metrics"])-1)
    sm_data   = Reference(ws2, min_col=4, min_row=4, max_row=4+len(d["sub_metrics"]))
    bar.add_data(sm_data, titles_from_data=True)
    bar.set_categories(sm_labels)
    ws2.add_chart(bar, "B13")

# ── Sheet 7: Evidence Index ────────────────────────────────────────────────────

ws3 = wb.create_sheet("Evidence Index")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 4
ws3.column_dimensions["B"].width = 10
ws3.column_dimensions["C"].width = 35
ws3.column_dimensions["D"].width = 20
ws3.column_dimensions["E"].width = 55

ws3.merge_cells("B2:E2")
ws3["B2"] = f"Evidence Index — {COMPANY}"
ws3["B2"].font = Font(name="Calibri", bold=True, size=13, color=WHITE)
ws3["B2"].fill = header_fill(DARK_BLUE)
ws3["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[2].height = 24

for col, val in [("B","Code"),("C","Sub-Metric"),("D","Source"),("E","Evidence")]:
    ws3[f"{col}4"] = val
    ws3[f"{col}4"].font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    ws3[f"{col}4"].fill = header_fill(MID_BLUE)
    ws3[f"{col}4"].border = thin_border()
    ws3[f"{col}4"].alignment = Alignment(horizontal="center", vertical="center")

sources = [
    ("SL01","Founder/Company Tenure","Janatics website / MCA","Incorporated 1991; brand established 1977 — 49-year operating history"),
    ("SL02","Director Profile","MCA public records","3 active directors; AGM Sept 2024; ROC Coimbatore filings current"),
    ("SL03","Governance Compliance","MCA / Zaubacorp","Annual filings current; AGM held within statutory deadline"),
    ("SL04","Debt Structure","ICRA Rationale Report","ICRA rated; stable outlook; healthy debt protection metrics; multiple reaffirmations"),
    ("SL05","Strategic Direction","Janatics website","Expanded to electric actuators, sensors, motors, robotics — Industry 4.0 pivot"),
    ("SL06","Global Presence","janaticspneumatics.com","42+ countries; subsidiaries USA and Germany; 200 global distribution partners"),
    ("SM01","Revenue Growth","ICRA / Tracxn","FY24: ₹507.9 Cr → FY25: ₹527.8 Cr; ~4% YoY vs 6.9% industry CAGR"),
    ("SM02","Revenue Scale","Tracxn / ICRA 2025","₹531 Cr FY2025; consistent ₹500 Cr+ revenue for 2 consecutive years"),
    ("SM03","Customer Diversity","Janatics website","17,000+ customers; 6+ industries: pharma, auto, packaging, food, medical, textile"),
    ("SM04","Distribution Network","janaticspneumatics.com","200 global distribution partners; 3 manufacturing plants for regional reach"),
    ("SM05","Brand Strength","AIA India / DSIR","Member of Automation Industry Association India; DSIR-recognized R&D division"),
    ("OSC01","Manufacturing Scale","janaticspneumatics.com","70,000 sqm Coimbatore HQ; additional Ahmedabad and Noida plants"),
    ("OSC02","Product Portfolio","ICRA Rationale","3,500+ distinct products across pneumatics, vacuum, hydraulics, didactics, robotics"),
    ("OSC03","R&D Capability","DSIR / Janatics website","DSIR-recognized R&D; CNC and casting; expanding to robotic systems"),
    ("OSC04","GST Compliance","ICRA / public records","Active GST entity; ICRA notes consistent regulatory compliance"),
    ("OSC05","Digitization","Janatics website","Industry 4.0 product line exists; internal ERP/digitization not publicly disclosed"),
    ("FIN01","Revenue FY2025","ICRA Rationale 2025","FY2025 revenue: ₹527.8 Cr (ICRA) / ₹531 Cr (Tracxn)"),
    ("FIN02","Revenue Growth","ICRA","4% YoY growth FY24→FY25; 3-yr CAGR not publicly available"),
    ("FIN03","EBITDA Trajectory","Tracxn","EBITDA CAGR -22% (1-year) — margin compression; specific % not disclosed"),
    ("FIN04","Credit Profile","ICRA","Stable outlook; healthy credit; multiple reaffirmations over 3+ years"),
    ("FIN05","Capital Structure","MCA / Zaubacorp","Auth. ₹30 Cr; paid-up ₹26.5 Cr; no external equity funding rounds identified"),
    ("FIN06","Employee Productivity","Tracxn","662 employees; ₹531 Cr revenue = ~₹80 lakh revenue per employee"),
    ("IC01","Market Size","IMARC Group 2024","India pneumatic equipment market $1.06B (2024) → $1.94B (2033)"),
    ("IC02","Industry CAGR","IMARC Group","India pneumatic equipment CAGR 6.9% (2025–2033)"),
    ("IC03","Competitive Position","Public research","Competes with SMC, Festo, Parker Hannifin; dominant domestic brand in mid-market"),
    ("IC04","Macro Tailwinds","Industry reports","Make in India, PLI for manufacturing, China+1 sourcing shifts — all favourable"),
    ("IC05","End-Market Mix","ICRA / Janatics","Pharma, auto, packaging, printing, food processing, medical equipment, textile — 7 sectors"),
]

for j, row in enumerate(sources, start=5):
    ws3.row_dimensions[j].height = 28
    for col, val in zip(["B","C","D","E"], row):
        cell = ws3[f"{col}{j}"]
        cell.value = val
        cell.font = Font(name="Calibri", size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = thin_border()
        if col == "B":
            cell.font = Font(name="Calibri", bold=True, size=9, color=MID_BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center")

# ── Save ───────────────────────────────────────────────────────────────────────

out = "/Users/laaksandy/Documents/Claude-testing/DDMR/output/Janatics_DDMR_Scorecard.xlsx"
wb.save(out)
print(f"Saved: {out}")
